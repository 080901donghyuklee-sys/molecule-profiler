import streamlit as st
import re
from collections import deque
import openpyxl
from rdkit import Chem
from rdkit.Chem import Descriptors


#1. 그래프 이론 기반 파싱 및 유틸리티

def tokenize_smiles(smiles):
    tokens = []
    i, L = 0, len(smiles)
    while i < L:
        if smiles[i] == '[':
            j = smiles.find(']', i)
            if j != -1:
                tokens.append(smiles[i:j+1])
                i = j + 1
            else:
                tokens.append(smiles[i])
                i += 1
        elif smiles[i:i+2] in ["Cl", "Br", "Na", "Ca", "Si"]:
            tokens.append(smiles[i:i+2])
            i += 2
        elif smiles[i] in ['(', ')', '=', '#', '%']:
            tokens.append(smiles[i])
            i += 1
        else:
            tokens.append(smiles[i])
            i += 1
    return tokens

def clean_substituent(sub):
    if not sub: return ""
    if not (sub.startswith('R') or sub.startswith('r')):
        sub = 'R' + sub

    digits = [c for c in sub if c.isdigit()]
    counts = {d: digits.count(d) for d in set(digits)}
    dangling = [d for d, c in counts.items() if c % 2 != 0]
    
    res = sub
    for d in dangling:
        res = re.sub(r'([A-Z][a-z]?|[a-z])' + str(d), 'R', res)
        
    res = re.sub(r'R+', 'R', res)      
    res = res.replace('R=R', 'R=')     
    res = res.replace('R#R', 'R#')
    res = res.replace('(R)', '')       
    res = res.replace('()', '')
    
    if res.startswith('=R'): res = 'R=' + res[2:]
    if res.startswith('#R'): res = 'R#' + res[2:]
    return res

def build_smiles_graph(smiles):
    num_double_at = smiles.count('@@')
    num_single_at = smiles.count('@') - (num_double_at * 2)
    stereo_counts = {
        '@@': num_double_at,
        '@': num_single_at,
        '/': smiles.count('/'),
        '\\': smiles.count('\\')
    }
    
    smiles = re.sub(r'\[([A-Za-z][a-z]?)@@?H?\]', r'\1', smiles)
    for sym in ['/', '\\']:
        smiles = smiles.replace(sym, '')

    atoms, adj, bonds = [], {}, {}
    ring_closures = {}
    ring_pairs = []
    
    i, L = 0, len(smiles)
    last_atom, branch_stack = None, []
    current_bond = ""
    
    while i < L:
        c = smiles[i]
        if c in ['=', '#']:
            current_bond = c; i += 1; continue
        if c == '(':
            branch_stack.append((last_atom, current_bond))
            current_bond = ""; i += 1; continue
        if c == ')':
            if branch_stack: last_atom, current_bond = branch_stack.pop()
            current_bond = ""; i += 1; continue
        if c == '[':
            j = smiles.find(']', i+1)
            token = smiles[i:j+1]
            idx = len(atoms)
            atoms.append(token)
            adj[idx] = []
            if last_atom is not None:
                adj[last_atom].append(idx); adj[idx].append(last_atom)
                bonds[(last_atom, idx)] = current_bond; bonds[(idx, last_atom)] = current_bond
            last_atom = idx; current_bond = ""; i = j + 1; continue
        if c.isalpha():
            token = c
            if (i+1 < L) and smiles[i+1].islower():
                two = smiles[i:i+2]
                if two in ["Cl", "Br", "Na", "Ca", "Si"]: token = two; i += 1
            idx = len(atoms)
            atoms.append(token)
            adj[idx] = []
            if last_atom is not None:
                adj[last_atom].append(idx); adj[idx].append(last_atom)
                bonds[(last_atom, idx)] = current_bond; bonds[(idx, last_atom)] = current_bond
            last_atom = idx; current_bond = ""; i += 1; continue
        if c.isdigit() or c == '%':
            if c == '%': num = int(smiles[i+1:i+3]); i += 3
            else: num = int(c); i += 1
            if last_atom is None: continue
            if num not in ring_closures:
                ring_closures[num] = (last_atom, current_bond)
            else:
                closing_atom, closing_bond = ring_closures[num]
                final_bond = current_bond if current_bond else closing_bond
                if closing_atom not in adj[last_atom]:
                    adj[last_atom].append(closing_atom); adj[closing_atom].append(last_atom)
                    bonds[(last_atom, closing_atom)] = final_bond; bonds[(closing_atom, last_atom)] = final_bond
                ring_pairs.append((num, closing_atom, last_atom))
                del ring_closures[num]
            current_bond = ""; continue
        if c == '.': last_atom = None; current_bond = ""; i += 1; continue
        i += 1
        
    return atoms, adj, bonds, ring_pairs, stereo_counts

def find_all_rings(atoms, adj, ring_pairs):
    rings = []
    for num, u, v in ring_pairs:
        visited, parent = {u}, {u: None}
        q = deque([u])
        found = False
        while q:
            curr = q.popleft()
            if curr == v: found = True; break
            for nxt in adj[curr]:
                if (curr == u and nxt == v) or (curr == v and nxt == u): continue
                if nxt not in visited:
                    visited.add(nxt); parent[nxt] = curr; q.append(nxt)
        if found:
            path = []
            curr = v
            while curr is not None:
                path.append(curr)
                curr = parent[curr]
            rings.append({"num": num, "atoms": set(path), "list": path[::-1]})
    rings.sort(key=lambda x: x["num"])
    return rings

def reconstruct_substituent_smiles(atoms, adj, bonds, start, avoid, original_rings):
    visited = set(avoid)
    tree_children = {i: [] for i in range(len(atoms))}
    back_edges = []
    
    def explore(u, p):
        visited.add(u)
        for v in adj[u]:
            if v == p or v in avoid: continue
            if v in visited:
                if u < v:
                    ring_num = next((r["num"] for r in original_rings if u in r["atoms"] and v in r["atoms"]), 99)
                    back_edges.append((u, v, ring_num))
            else:
                tree_children[u].append(v); explore(v, u)
                
    explore(start, None)
    node_closures = {i: [] for i in range(len(atoms))}
    for u, v, r_num in back_edges:
        bond = bonds[(u, v)]
        node_closures[u].append((r_num, bond))
        node_closures[v].append((r_num, ""))
        
    def build_str(u):
        res = atoms[u]
        for r, b in node_closures[u]: res += f"{b}%{r}" if r >= 10 else f"{b}{r}"
        children = tree_children[u]
        for i, child in enumerate(children):
            bond = bonds[(u, child)]
            child_str = bond + build_str(child)
            res += f"({child_str})" if i < len(children) - 1 else child_str
        return res
    return build_str(start)

def advanced_split_substituent(sub, bond_counts):
    if not sub: return []
    pure_sub = sub[1:] if sub[0] in ['R', 'r'] else sub
    tokens = tokenize_smiles(pure_sub)
    results, current_piece = [], []
    
    i, L = 0, len(tokens)
    while i < L:
        t = tokens[i]
        if t in ['O', 'o', 'N', 'n', 'S', 's']:
            bond_prefix = current_piece.pop() if current_piece and current_piece[-1] in ['=', '#'] else ""
            if current_piece:
                results.append(clean_substituent('R' + ''.join(current_piece)))
                current_piece = []
            
            bond_type = None
            if t in ['O', 'o']:
                is_ester = (len(results) > 0 and ('=O' in results[-1] or 'O=' in results[-1])) or bond_prefix == '='
                if i + 3 < L and tokens[i+1] == '(' and tokens[i+2] == '=' and tokens[i+3] in ['O', 'o']: is_ester = True
                bond_type = "Ester" if is_ester else "Ether"
            elif t in ['N', 'n']: bond_type = "Amine"
            elif t in ['S', 's']: bond_type = "Sulfide"
                
            if bond_type: bond_counts[bond_type] = bond_counts.get(bond_type, 0) + 1
            if bond_prefix: current_piece.append(bond_prefix)
            current_piece.append(t)
            i += 1; continue
        elif t == '(':
            if current_piece:
                results.append(clean_substituent('R' + ''.join(current_piece)))
                current_piece = []
            depth, j = 1, i + 1
            branch_tokens = []
            while j < L and depth > 0:
                if tokens[j] == '(': depth += 1
                elif tokens[j] == ')': depth -= 1
                if depth > 0: branch_tokens.append(tokens[j])
                j += 1
            if branch_tokens: results.append(clean_substituent('R' + ''.join(branch_tokens)))
            i = j; continue
        elif t == ')': i += 1; continue
        else: current_piece.append(t); i += 1
            
    if current_piece: results.append(clean_substituent('R' + ''.join(current_piece)))
    return results

def run_molecule_analysis_pipeline(smiles, functional_group_db, score_db, rdkit_res=None):
    atoms, adj, bonds, ring_pairs, stereo_counts = build_smiles_graph(smiles)
    
    #파서의 토큰 기반으로 탄소 개수 계산
    total_carbons = sum(1 for a in atoms if a.upper() == 'C')
    
    all_rings = find_all_rings(atoms, adj, ring_pairs)
    original_rings_formatted = [[r["num"]] + [atoms[k] for k in r["list"]] for r in all_rings]
    
    ring_systems = []
    for r in all_rings:
        merged = False
        for sys in ring_systems:
            if sys & r["atoms"]:
                sys.update(r["atoms"]); merged = True; break
        if not merged: ring_systems.append(set(r["atoms"]))
        
    changed = True
    while changed:
        changed = False
        for i in range(len(ring_systems)):
            for j in range(i+1, len(ring_systems)):
                if ring_systems[i] & ring_systems[j]:
                    ring_systems[i].update(ring_systems[j]); del ring_systems[j]
                    changed = True; break
            if changed: break
            
    ring_systems.sort(key=lambda s: min(s) if s else 0)
    main_system = ring_systems[0] if ring_systems else set()
    
    substituents_data = []
    visited_sub_atoms = set()
    
    for u in main_system:
        for v in adj[u]:
            if v not in main_system and v not in visited_sub_atoms:
                sub_atoms = set()
                q = deque([v])
                sub_atoms.add(v)
                while q:
                    curr = q.popleft()
                    for nxt in adj[curr]:
                        if nxt not in main_system and nxt not in sub_atoms:
                            sub_atoms.add(nxt); q.append(nxt)
                visited_sub_atoms.update(sub_atoms)
                
                contained_rings = [r for r in all_rings if r["atoms"].issubset(sub_atoms)]
                bond_to_core = bonds[(u, v)]
                sub_smiles = reconstruct_substituent_smiles(atoms, adj, bonds, v, main_system, all_rings)
                substituents_data.append({"smiles": clean_substituent(f"R{bond_to_core}{sub_smiles}"), "rings": contained_rings})
                
    sub_ring_list, count_table, alkyl_records, branch_list = [], {}, [], []
    bond_counts = {"Ester": 0, "Ether": 0, "Amine": 0, "Sulfide": 0}
    db_norm = {k.replace('r', 'R'): v for k, v in functional_group_db.items()}
    
    for sub_info in substituents_data:
        sub = sub_info["smiles"]
        if sub_info["rings"]:
            for r in sub_info["rings"]:
                formatted_r = [r["num"]] + [atoms[k] for k in r["list"]]
                if formatted_r not in sub_ring_list: sub_ring_list.append(formatted_r)
            continue
            
        task_list = deque([sub])
        while task_list:
            curr_sub = task_list.popleft()
            query = curr_sub.replace('r', 'R')
            rev_query = query[::-1]
            
            matched_group = db_norm.get(query) or db_norm.get(rev_query)
            if matched_group:
                count_table[matched_group] = count_table.get(matched_group, 0) + 1
                continue
                
            if all(char in "CH()rR" for char in curr_sub):
                carbon_count = curr_sub.count('C') + curr_sub.count('c')
                alkyl_records.append({"substituent": curr_sub, "carbon_count": carbon_count})
                continue
                
            fork_pieces = advanced_split_substituent(curr_sub, bond_counts)
            if len(fork_pieces) > 1 or (fork_pieces and fork_pieces[0] != curr_sub):
                task_list.extend(fork_pieces)
                branch_list.append({"parent": curr_sub, "split_branches": fork_pieces})
            else:
                branch_list.append({"parent": curr_sub, "note": "Unresolvable chain"})
                
    res = {
        "original_rings": original_rings_formatted,
        "substituent_rings": sub_ring_list,
        "functional_groups_count": count_table,
        "alkyl_groups": alkyl_records,
        "branches": branch_list,
        "stereo_counts": stereo_counts,
        "special_bond_counts": bond_counts
    }
    
    if score_db: 
        res["property_scores"] = calculate_scores(
            count_table, score_db, original_rings_formatted, rdkit_res, total_carbons
        )
    return res


#2. 파일 로드 및 한글 변환 스코어링 모듈

@st.cache_data
def load_functional_group_db(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        return {str(row[0]).strip(): str(row[1]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    except:
        return {}

@st.cache_data
def load_score_table(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        score_db = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                score_db[str(row[0]).strip()] = {
                    "group_name": row[1], "polarity": row[2] or 0, "hydrophobicity": row[3] or 0,
                    "hbond_donor": row[4] or 0, "hbond_acceptor": row[5] or 0, "bbb": row[6] or 0,
                    "acidity": row[7] or 0, "basicity": row[8] or 0, "steric": row[9] or 0
                }
        return score_db
    except:
        return {}

def calculate_scores(functional_groups_count, score_db, original_rings=None, rdkit_res=None, total_carbons=0):

    total = {
        "극성": 0,
        "소수성": 0,
        "수소결합 주개": 0,
        "수소결합 받개": 0,
        "혈뇌장벽 투과성": 0,
        "산도": 0,
        "염기도": 0,
        "입체장애": 0
    }
    
    # 엑셀의 영어 행 키와 매칭할 한글 키 사전 구축
    key_map = {
        "polarity": "극성",
        "hydrophobicity": "소수성",
        "hbond_donor": "수소결합 주개",
        "hbond_acceptor": "수소결합 받개",
        "bbb": "혈뇌장벽 투과성",
        "acidity": "산도",
        "basicity": "염기도",
        "steric": "입체장애"
    }
    
    # 1. 작용기 점수 누적 (엑셀 데이터 기반)
    for group_name, count in functional_groups_count.items():
        for pattern, data in score_db.items():
            if data["group_name"] == group_name:
                for eng_key, kor_key in key_map.items():
                    total[kor_key] += data[eng_key] * count
                break
                
    # 2. 고리 골격 기반 원자 가산
    if original_rings:
        for ring in original_rings:
            atoms_in_ring = ring[1:]
            for atom in atoms_in_ring:
                atom_upper = str(atom).upper()
                if 'N' in atom_upper: 
                    total["극성"] += 1
                    total["수소결합 받개"] += 1
                    total["염기도"] += 1
                    total["혈뇌장벽 투과성"] -= 1
                elif 'O' in atom_upper: 
                    total["극성"] += 1
                    total["수소결합 받개"] += 1
                    total["산도"] += 1
                    total["혈뇌장벽 투과성"] -= 1
                elif 'C' in atom_upper or 'S' in atom_upper: 
                    total["소수성"] += 1
                    total["입체장애"] += 1
                    
    # 3. [탄소 보너스 엔진] 총 탄소 개수에 따른 어드밴티지 가산
    total["소수성"] += total_carbons // 4
    total["입체장애"] += total_carbons // 5
                    
    # 4. RDKit 분자량 조건 연산 및 BBB 페널티 상쇄 보정
    if rdkit_res and "mw" in rdkit_res:
        mw = rdkit_res["mw"]
        total["입체장애"] += int(mw // 100)
        total["극성"] += int(mw // 150)
        
        if mw > 500:
            
            if total_carbons >= 35:
                total["혈뇌장벽 투과성"] += 0          
                total["소수성"] += 4           
            else:
                total["혈뇌장벽 투과성"] -= 2
                total["소수성"] += 1
        elif mw < 350:
            total["혈뇌장벽 투과성"] += 1
            
    return total


#3. RDKit 기초 특성 프로파일링 

def analyze_rdkit_properties(smiles):
    mol = Chem.MolFromSmiles(smiles)
    if mol is None: return None
    
    return {
        "mw": Descriptors.MolWt(mol),
        "logp": Descriptors.MolLogP(mol),
        "hbd": Descriptors.NumHDonors(mol),
        "hba": Descriptors.NumHAcceptors(mol),
        "rings": []
    }


#4. 약물 계열 예측 모듈 

def predict_drug_class(smiles, rdkit_res, custom_res):
    predicted_classes = []
    mol = Chem.MolFromSmiles(smiles)
    
    ring_sizes = [len(r) - 1 for r in custom_res.get("original_rings", [])]
    ring_elements = [r[1:] for r in custom_res.get("original_rings", [])]
    fgs = custom_res.get("functional_groups_count", {})
    special_bonds = custom_res.get("special_bond_counts", {})
    
    def has_smarts(target_mol, smarts_str):
        if target_mol is None: return False
        query = Chem.MolFromSmarts(smarts_str)
        if query is None: return False
        return target_mol.HasSubstructMatch(query)

    if 4 in ring_sizes and (5 in ring_sizes or 6 in ring_sizes):
        four_membered_rings = [re for rs, re in zip(ring_sizes, ring_elements) if rs == 4]
        if any('N' in r or 'n' in r for r in four_membered_rings):
            predicted_classes.append("베타-락탐 계열 항생제 (Beta-lactam Antibiotics)")
            
    if 6 in ring_sizes and 7 in ring_sizes:
        seven_membered_rings = [re for rs, re in zip(ring_sizes, ring_elements) if rs == 7]
        if any(r.count('N') + r.count('n') >= 2 for r in seven_membered_rings):
            predicted_classes.append("벤조디아제핀 계열 신경안정제 (Benzodiazepines)")
            
    if ring_sizes.count(6) >= 3 and ring_sizes.count(5) >= 1:
        if has_smarts(mol, "C1CCC2C(C1)CCC3C2CCC4C3CCC4"):
            predicted_classes.append("스테로이드 계열 호르몬 (Steroids)")
            
    if ring_sizes.count(6) >= 1 and special_bonds.get("Ester", 0) >= 2:
        predicted_classes.append("디히드로피리딘 계열 고혈압약 (Dihydropyridines - CCB)")
        
    has_sulfonamide = any("sulfonamide" in k.lower() for k in fgs.keys())
    if has_sulfonamide or has_smarts(mol, "S(=O)(=O)N"):
        predicted_classes.append("설폰아미드 계열 항균/이뇨제 (Sulfonamides)")
        
    if has_smarts(mol, "O=C(O)CC(O)CC(O)") or has_smarts(mol, "O=C1CC(O)CC(=O)O1"):
        predicted_classes.append("스태틴 계열 고지혈증 치료제 (Statins)")
        
    has_f = any("fluor" in k.lower() or k == "F" for k in fgs.keys()) or ('F' in smiles)
    has_carboxyl = any("carboxyl" in k.lower() or k == "COOH" for k in fgs.keys()) or ('C(=O)O' in smiles)
    if ring_sizes.count(6) >= 2 and has_f and has_carboxyl:
        predicted_classes.append("플루오로퀴놀론 계열 항생제 (Fluoroquinolones)")
        
    if ring_sizes.count(6) >= 1 and ring_sizes.count(5) >= 1:
        if has_smarts(mol, "c1ccc2c(c1)c(c[nH]2)CCN"):
            predicted_classes.append("트리프탄 계열 편두통 치료제 (Triptans)")
            
    if ring_sizes.count(6) >= 1:
        six_membered_rings = [re for rs, re in zip(ring_sizes, ring_elements) if rs == 6]
        has_n2 = any(r.count('N') + r.count('n') >= 2 for r in six_membered_rings)
        carbonyl_cnt = sum(v for k, v in fgs.items() if "carbonyl" in k.lower() or "ketone" in k.lower())
        if has_n2 and (carbonyl_cnt >= 2 or smiles.count('=O') >= 3):
            predicted_classes.append("바비튜레이트 계열 중추신경억제제 (Barbiturates)")
            
    if ring_sizes.count(6) >= 3:
        if has_smarts(mol, "c1ccc2c(c1)Sc3ccccc3N2"):
            predicted_classes.append("페노티아진 계열 항정신병약 (Phenothiazines)")
            
    if rdkit_res.get("mw", 0) > 700 and any(size >= 12 for size in ring_sizes):
        predicted_classes.append("리파마이신 계열 (Rifamycins)")
    elif has_smarts(mol, "O=CNC2=C(O)c3c(C2=O)cccc3"):
        predicted_classes.append("리파마이신 계열 (Rifamycins)")
            
    if not predicted_classes:
        return ["조건에 부합하는 특징이 없어 약물 계열 예측이 불가합니다."]
        
    return list(set(predicted_classes))



#5. Streamlit 웹 애플리케이션 프론트엔드 UI

st.set_page_config(page_title="Ultimate Molecule Profiler", layout="wide")

def main():
    st.title("Ultimate Molecule Profiler")
    st.markdown("그래프 이론 및 RDKit 하이브리드 화합물 분석 대시보드")
    
    EXCEL_FILE = "functional_group_score_table.xlsx"
    fg_db = load_functional_group_db(EXCEL_FILE)
    score_db = load_score_table(EXCEL_FILE)
    
    if not fg_db:
        st.warning("'functional_group_score_table.xlsx' 파일을 찾을 수 없어 작용기 매칭 및 점수 산출이 제한됩니다.")
    
    st.markdown("---")
    
    smiles_input = st.text_input("**SMILES 문자열을 입력하세요:**", placeholder="예: C1=CC=CC=C1")
    analyze_btn = st.button("분석 실행", type="primary")
    
    if analyze_btn and smiles_input:
        with st.spinner("분자를 프로파일링 하는 중..."):
            
            rdkit_res = analyze_rdkit_properties(smiles_input.strip())
            if rdkit_res is None:
                st.error("유효하지 않은 SMILES 문자열입니다. 다시 입력해주세요.")
                return
                
            custom_res = run_molecule_analysis_pipeline(smiles_input.strip(), fg_db, score_db, rdkit_res)
            predicted_drug_classes = predict_drug_class(smiles_input.strip(), rdkit_res, custom_res)
            
            st.success("분석 완료")
            
            
            # 1) 기초 물리화학적 특성
            
            st.subheader(" PART 1. 기초 물리화학적 특성 (RDKit)")
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("분자량(MW)", f"{rdkit_res['mw']:.2f} g/mol")
            col2.metric("지질친화성(LogP)", f"{rdkit_res['logp']:.2f}")
            col3.metric("수소결합 주개(HBD)", rdkit_res['hbd'])
            col4.metric("수소결합 받개(HBA)", rdkit_res['hba'])
            
            st.divider()
            
           
            # 2) 고리 시스템 프로파일링 (SMILES 고리 넘버 기준)
           
            st.subheader("PART 2. 고리 시스템 프로파일링")
            if custom_res.get('original_rings'):
                for ring_info in custom_res['original_rings']:
                    ring_num = ring_info[0]
                    atoms_in_ring = ring_info[1:]
                    ring_size = len(atoms_in_ring)
                    
                    atom_upper = [str(a).upper() for a in atoms_in_ring]
                    c_count = atom_upper.count('C')
                    n_count = atom_upper.count('N')
                    o_count = atom_upper.count('O')
                    s_count = atom_upper.count('S')
                    hetero_count = n_count + o_count + s_count
                    
                    if ring_size == 3 and c_count == 3: desc = "사이클로프로판형 (대사 안정성 및 소수성 강화)"
                    elif ring_size == 5 and n_count == 1: desc = "피롤/피롤리딘형 헤테로환 (의약품 핵심 잔가지)"
                    elif ring_size == 6 and c_count == 6: desc = "벤젠/사이클로헥산형 (소수성 코어 뼈대)"
                    elif ring_size == 6 and n_count == 1: desc = "피리딘/피페리딘형 헤테로환 (염기성 및 수소결합 능력 부여)"
                    else: desc = f"{ring_size}원환 구조 (C:{c_count}, Hetero:{hetero_count})"
                        
                    st.write(f"- **[{ring_num}번 고리]**: {desc}")
            else:
                st.info("고리가 없는 선형/사슬형 구조입니다.")
                
            st.divider()
            
          
            # 3) 정밀 입체/작용기/결합 분석
       
            st.subheader("PART 3. 정밀 입체/작용기/결합 분석 (Graph Theory)")
            col_a, col_b, col_c = st.columns(3)
            
            with col_a:
                st.markdown("**입체화학 (Chiral Centers)**")
                has_stereo = False
                for k, v in custom_res["stereo_counts"].items():
                    if v > 0:
                        st.write(f"- {k} : {v}개")
                        has_stereo = True
                if not has_stereo: st.write("- (없음)")
                
            with col_b:
                st.markdown("**발견된 기지 작용기**")
                if custom_res["functional_groups_count"]:
                    for k, v in custom_res["functional_groups_count"].items():
                        st.write(f"- {k} : {v}개")
                else: st.write("- (없음)")
                
            with col_c:
                st.markdown("**발견된 특수 결합 (Backbone)**")
                has_bonds = False
                for k, v in custom_res["special_bond_counts"].items():
                    if v > 0:
                        st.write(f"- {k} : {v}개")
                        has_bonds = True
                if not has_bonds: st.write("- (없음)")
                
            st.divider()
            
         
            # 4) 작용기 기반 환산 점수 
         
            if custom_res.get("property_scores"):
                st.subheader("PART 4. 작용기 기반 환산 점수")
                score_cols = st.columns(4)
                scores = list(custom_res["property_scores"].items())
                for i, (k, v) in enumerate(scores):
                    
                    score_cols[i % 4].metric(k, v)
                st.divider()
                
           
            # 5) 구조 기반 약물 계열 예측
           
            st.subheader("PART 5. 구조 기반 약물 계열 예측")
            for p_class in predicted_drug_classes:
                if "조건에 부합하는 특징이 없어" in p_class:
                    st.warning(p_class)
                else:
                    st.info(f"**{p_class}**")

if __name__ == "__main__":
    main()
